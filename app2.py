# app.py — Schedule Quality Validator | JESA Enterprise Edition

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import tempfile, os, io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from xer_parser.xer_parser          import XERParser
from metrics.open_ends               import check_open_ends
from metrics.leads                   import check_leads
from metrics.lags                    import check_lags
from metrics.relationship_types      import check_relationship_types
from metrics.hard_constraints        import check_hard_constraints
from metrics.high_float              import check_high_float
from metrics.negative_float          import check_negative_float
from metrics.long_duration           import check_long_duration
from metrics.invalid_dates           import check_invalid_dates
from metrics.resources               import check_resources
from metrics.remaining_metrics       import (check_logic_density,
                                             check_missed_activities,
                                             check_summary_tasks,
                                             check_critical_path)
from metrics.redundant_relationships import check_redundant_relationships
from engineering.mechanical_checks   import check_mechanical_sequence
from engineering.energy_checks       import check_energy_dependencies

# ══════════════════════════════════════════════════════════
# PAGE CONFIG
# ══════════════════════════════════════════════════════════
st.set_page_config(
    page_title="JESA — Schedule Quality Validator",
    page_icon="📅",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ══════════════════════════════════════════════════════════
# ENTERPRISE CSS
# ══════════════════════════════════════════════════════════
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

    * { font-family: 'Inter', sans-serif !important; }

    .stApp { background-color: #f5f7fa; }

    /* ── Sidebar ── */
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0a1628 0%, #0d2151 60%, #0a1628 100%);
        border-right: 1px solid #1e3a6e;
    }
    section[data-testid="stSidebar"] * { color: #e8edf5 !important; }
    section[data-testid="stSidebar"] hr { border-color: #1e3a6e !important; }

    /* ── Top Banner ── */
    .enterprise-header {
        background: linear-gradient(135deg, #0a1628 0%, #0d2151 50%, #1a3a7e 100%);
        padding: 24px 36px;
        border-radius: 14px;
        margin-bottom: 24px;
        display: flex;
        align-items: center;
        box-shadow: 0 4px 24px rgba(10,22,40,0.18);
        border: 1px solid #1e3a6e;
    }
    .header-title {
        font-size: 24px; font-weight: 800;
        color: #ffffff; margin: 0; letter-spacing: 0.3px;
    }
    .header-subtitle {
        font-size: 12px; color: #90afd4; margin-top: 4px;
        letter-spacing: 0.8px; text-transform: uppercase;
    }
    .header-badge {
        background: rgba(255,255,255,0.1);
        border: 1px solid rgba(255,255,255,0.2);
        color: #90afd4 !important;
        padding: 4px 12px; border-radius: 20px;
        font-size: 11px; font-weight: 600;
        text-transform: uppercase; letter-spacing: 1px;
        margin-left: 12px; display: inline-block;
    }

    /* ── Score Banner ── */
    .score-banner {
        border-radius: 14px; padding: 28px 36px;
        margin-bottom: 24px;
        box-shadow: 0 4px 20px rgba(0,0,0,0.08);
        border: 1px solid rgba(255,255,255,0.6);
    }
    .score-number {
        font-size: 72px; font-weight: 800; line-height: 1;
        margin: 0;
    }
    .score-label {
        font-size: 13px; font-weight: 600;
        text-transform: uppercase; letter-spacing: 1px;
        opacity: 0.7; margin-top: 4px;
    }
    .score-grade {
        font-size: 22px; font-weight: 700; margin-top: 8px;
    }

    /* ── Metric Cards ── */
    .metric-card {
        background: #ffffff;
        border-radius: 12px;
        padding: 18px 20px;
        margin-bottom: 12px;
        box-shadow: 0 2px 12px rgba(0,0,0,0.06);
        border-left: 5px solid #ccc;
        transition: transform 0.15s ease;
    }
    .metric-card:hover { transform: translateY(-2px); }
    .metric-card.pass  { border-left-color: #00b894; }
    .metric-card.fail  { border-left-color: #d63031; }

    .metric-card-title {
        font-size: 13px; font-weight: 700; color: #2d3436;
        text-transform: uppercase; letter-spacing: 0.5px;
    }
    .metric-card-stat {
        font-size: 26px; font-weight: 800;
        margin: 4px 0;
    }
    .metric-card-sub {
        font-size: 11px; color: #636e72; font-weight: 500;
    }
    .badge-pass {
        background: #00b894; color: white;
        padding: 3px 10px; border-radius: 20px;
        font-size: 11px; font-weight: 700;
        text-transform: uppercase; letter-spacing: 0.5px;
    }
    .badge-fail {
        background: #d63031; color: white;
        padding: 3px 10px; border-radius: 20px;
        font-size: 11px; font-weight: 700;
        text-transform: uppercase; letter-spacing: 0.5px;
    }

    /* ── Section Title ── */
    .section-title {
        font-size: 13px; font-weight: 700; color: #636e72;
        text-transform: uppercase; letter-spacing: 1.2px;
        margin: 28px 0 14px 0;
        padding-bottom: 8px;
        border-bottom: 1px solid #dfe6e9;
    }

    /* ── KPI Box ── */
    .kpi-box {
        background: #ffffff;
        border-radius: 12px; padding: 20px 24px;
        box-shadow: 0 2px 12px rgba(0,0,0,0.06);
        text-align: center;
        border: 1px solid #f0f3f7;
    }
    .kpi-value {
        font-size: 38px; font-weight: 800; color: #0d2151;
    }
    .kpi-label {
        font-size: 11px; font-weight: 600; color: #636e72;
        text-transform: uppercase; letter-spacing: 1px;
        margin-top: 4px;
    }

    /* ── File Tab Header ── */
    .file-header {
        background: #ffffff; border-radius: 10px;
        padding: 14px 20px; margin-bottom: 20px;
        border: 1px solid #dfe6e9;
        box-shadow: 0 2px 8px rgba(0,0,0,0.04);
        display: flex; align-items: center; gap: 12px;
    }
    .file-name { font-size: 16px; font-weight: 700; color: #0d2151; }
    .file-sub  { font-size: 12px; color: #636e72; }

    /* ── Tabs ── */
    .stTabs [data-baseweb="tab-list"] {
        background: #ffffff; border-radius: 10px;
        border: 1px solid #dfe6e9; padding: 4px; gap: 4px;
    }
    .stTabs [data-baseweb="tab"] {
        color: #636e72 !important; font-weight: 600;
        font-size: 13px; border-radius: 8px !important;
        padding: 8px 18px !important;
    }
    .stTabs [aria-selected="true"] {
        background: #0d2151 !important;
        color: #ffffff !important;
    }

    /* ── Tables ── */
    .stDataFrame { border-radius: 10px; overflow: hidden; }
    .stDataFrame thead th {
        background: #0d2151 !important;
        color: white !important; font-weight: 600 !important;
    }

    /* ── Violations Expander ── */
    details {
        background: #ffffff !important;
        border: 1px solid #dfe6e9 !important;
        border-radius: 10px !important;
        margin-bottom: 8px !important;
    }
    details summary {
        font-weight: 600 !important; font-size: 13px !important;
        padding: 12px 16px !important;
    }

    /* ── Upload Zone ── */
    .stFileUploader {
        background: #ffffff;
        border: 2px dashed #0d2151;
        border-radius: 12px;
    }

    /* ── Download Buttons ── */
    .stDownloadButton button {
        background: #0d2151 !important; color: #ffffff !important;
        border-radius: 8px !important; font-weight: 600 !important;
        border: none !important; padding: 10px 24px !important;
        font-size: 13px !important;
    }
    .stDownloadButton button:hover { background: #1a3a7e !important; }

    /* ── Info/Success/Warning ── */
    .stAlert { border-radius: 10px !important; }

    #MainMenu {visibility: hidden;}
    footer    {visibility: hidden;}
    header    {visibility: hidden;}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════
with st.sidebar:
    st.image("jesa_logo.png", width=140)
    st.markdown("---")

    st.markdown("### 📋 Reference Standard")
    st.info("**DCMA 14-Point Assessment**\nAll thresholds follow official DCMA guidelines.")

    st.markdown("---")
    st.markdown("### 📐 DCMA Thresholds")
    st.markdown("""
| Metric | Limit |
|---|---|
| Open Ends | < 5% |
| Leads | 0% |
| Lags | 0% |
| SF Relationships | 0% |
| Hard Constraints | < 5% |
| High Float | < 5% |
| Negative Float | 0% |
| Long Duration | < 5% |
| Invalid Dates | 0% |
| Resources | 0% |
| Logic Density | > 90% |
| Missed Activities | 0% |
| Critical Path | Defined |
| Summary Tasks | 0% |
    """)

    st.markdown("---")
    st.markdown("""
    **Schedule Quality Validator**
    Version 2.0 — PFE 2026

    Automated Primavera P6 XER
    quality assessment tool.

    © 2026 JESA — The Solution Company for Africa
    """)


# ══════════════════════════════════════════════════════════
# HEADER
# ══════════════════════════════════════════════════════════
col_logo, col_title = st.columns([1, 6])
with col_logo:
    st.image("jesa_logo.png", width=110)
with col_title:
    st.markdown("""
    <div class="enterprise-header">
        <div>
            <p class="header-title">Schedule Quality Validation Platform</p>
            <p class="header-subtitle">
                Primavera P6 XER &nbsp;·&nbsp; DCMA 14-Point Assessment
                &nbsp;·&nbsp; EPC Engineering Validation &nbsp;·&nbsp; Multi-File Analysis
            </p>
            <span class="header-badge">Enterprise Edition</span>
            <span class="header-badge">v2.0</span>
        </div>
    </div>
    """, unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════
# FILE UPLOAD
# ══════════════════════════════════════════════════════════
uploaded_files = st.file_uploader(
    "📂 Upload Primavera P6 XER Schedule Files",
    type=["xer"],
    accept_multiple_files=True,
    help="Upload one or more Primavera P6 XER files for automated quality assessment"
)

if not uploaded_files:
    st.markdown("""
    <div style="text-align:center; padding:70px 40px; background:#ffffff;
                border-radius:16px; border: 2px dashed #0d2151; margin-top:10px;">
        <div style="font-size:48px; margin-bottom:16px;">📂</div>
        <h2 style="color:#0d2151; font-weight:800; margin:0;">Upload Your Schedule Files</h2>
        <p style="color:#636e72; margin-top:10px; font-size:14px;">
            Drop one or more Primavera P6 <strong>.xer</strong> files above
            to begin the automated quality assessment.
        </p>
        <p style="color:#b2bec3; font-size:12px; margin-top:8px;">
            Supports all Primavera P6 XER versions &nbsp;·&nbsp;
            Multi-file batch analysis &nbsp;·&nbsp;
            DCMA 14-Point + Engineering Validation
        </p>
    </div>
    """, unsafe_allow_html=True)
    st.stop()


# ══════════════════════════════════════════════════════════
# PARSE ALL FILES
# ══════════════════════════════════════════════════════════
all_file_data = []

with st.spinner(f"⏳ Parsing and analysing {len(uploaded_files)} file(s)..."):
    for uploaded_file in uploaded_files:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xer") as tmp:
            tmp.write(uploaded_file.getvalue())
            tmp_path = tmp.name
        try:
            parser = XERParser(tmp_path)
            parser.parse()
        except Exception as e:
            st.error(f"❌ Failed to parse **{uploaded_file.name}**: {str(e)}")
            os.unlink(tmp_path)
            continue
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

        activities    = parser.activities
        relationships = parser.relationships

        proj_name     = "N/A"
        data_date_str = None
        proj_df       = parser.dataframes.get('PROJECT')
        if proj_df is not None:
            if 'proj_short_name' in proj_df.columns:
                proj_name = str(proj_df['proj_short_name'].iloc[0])
            for col in ['last_recalc_date', 'plan_start_date', 'scd_end_date']:
                if col in proj_df.columns:
                    val = str(proj_df[col].iloc[0])
                    if val and val != 'nan':
                        data_date_str = val[:10]
                        break

        results = [
            check_open_ends(activities),
            check_leads(activities, relationships),
            check_lags(activities, relationships),
            check_relationship_types(activities, relationships),
            check_hard_constraints(activities),
            check_high_float(activities),
            check_negative_float(activities),
            check_long_duration(activities),
            check_invalid_dates(activities),
            check_resources(activities),
            check_logic_density(activities),
            check_missed_activities(activities, data_date_str),
            check_summary_tasks(activities),
            check_critical_path(activities),
            check_redundant_relationships(activities, relationships),
        ]
        eng_results = [
            check_mechanical_sequence(activities, relationships),
            check_energy_dependencies(activities, relationships),
        ]

        passed_count = sum(1 for r in results if r['status'] == "PASS")
        failed_count = len(results) - passed_count
        score        = round((passed_count / len(results)) * 100)

        if score >= 80:
            grade = "GOOD"; grade_icon = "✅"; score_color = "#00b894"; bg_color = "#f0fff8"
        elif score >= 60:
            grade = "FAIR"; grade_icon = "⚠️"; score_color = "#f39c12"; bg_color = "#fffbf0"
        else:
            grade = "POOR"; grade_icon = "❌"; score_color = "#d63031"; bg_color = "#fff5f5"

        all_file_data.append({
            "filename"     : uploaded_file.name,
            "activities"   : activities,
            "relationships": relationships,
            "proj_name"    : proj_name,
            "data_date"    : data_date_str,
            "results"      : results,
            "eng_results"  : eng_results,
            "passed"       : passed_count,
            "failed"       : failed_count,
            "score"        : score,
            "score_color"  : score_color,
            "bg_color"     : bg_color,
            "grade"        : grade,
            "grade_icon"   : grade_icon,
        })

if not all_file_data:
    st.error("No files could be parsed. Please check your XER files.")
    st.stop()

st.success(f"✅ **{len(all_file_data)} file(s)** analysed successfully.")


# ══════════════════════════════════════════════════════════
# MULTI-FILE COMPARISON
# ══════════════════════════════════════════════════════════
if len(all_file_data) > 1:
    st.markdown('<p class="section-title">📊 Multi-File Comparison</p>', unsafe_allow_html=True)

    cmp_cols = st.columns(len(all_file_data))
    for i, fd in enumerate(all_file_data):
        with cmp_cols[i]:
            st.markdown(f"""
            <div class="kpi-box" style="border-top: 4px solid {fd['score_color']};">
                <div class="kpi-value" style="color:{fd['score_color']};">{fd['score']}%</div>
                <div class="kpi-label">{fd['filename'][:28]}</div>
                <div style="margin-top:8px; font-size:13px; font-weight:700; color:{fd['score_color']};">
                    {fd['grade_icon']} {fd['grade']}
                </div>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    fig_cmp = go.Figure()
    fig_cmp.add_trace(go.Bar(
        x=[fd["filename"][:22] for fd in all_file_data],
        y=[fd["score"] for fd in all_file_data],
        marker_color=[fd["score_color"] for fd in all_file_data],
        text=[f"{fd['score']}%" for fd in all_file_data],
        textposition="outside", textfont=dict(size=13, color="#2d3436")
    ))
    fig_cmp.update_layout(
        paper_bgcolor="#ffffff", plot_bgcolor="#ffffff",
        height=280, margin=dict(t=20, b=60, l=40, r=20),
        yaxis=dict(range=[0,115], title="Quality Score (%)",
                   gridcolor="#f0f3f7", tickfont=dict(color="#636e72")),
        xaxis=dict(tickangle=-15, tickfont=dict(color="#2d3436", size=12)),
        showlegend=False
    )
    st.plotly_chart(fig_cmp, use_container_width=True)


# ══════════════════════════════════════════════════════════
# EXCEL EXPORT HELPER  (defined once, outside loop)
# ══════════════════════════════════════════════════════════
def generate_excel(fd):
    wb          = Workbook()
    green_fill  = PatternFill("solid", fgColor="C6EFCE")
    red_fill    = PatternFill("solid", fgColor="FFC7CE")
    navy_fill   = PatternFill("solid", fgColor="0D2151")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    def style_header(ws, row_num, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = navy_fill; cell.font = header_font
            cell.alignment = center; cell.border = thin

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    # Sheet 1 — Executive Summary
    ws1 = wb.active; ws1.title = "Executive Summary"
    ws1.merge_cells("A1:G1")
    ws1["A1"] = f"SCHEDULE QUALITY REPORT — {fd['proj_name']} — {fd['filename']}"
    ws1["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws1["A1"].fill = navy_fill; ws1["A1"].alignment = center
    ws1.append([])
    ws1.append(["Project", fd['proj_name'], "", "Score", f"{fd['score']}%",
                "Grade", f"{fd['grade_icon']} {fd['grade']}"])
    ws1.append(["Activities", len(fd['activities']), "", "Relationships",
                len(fd['relationships']), "Data Date", fd['data_date'] or "N/A"])
    ws1.append([])
    headers = ["#", "Metric", "Status", "Violations", "Total", "Rate", "Threshold"]
    ws1.append(headers); style_header(ws1, ws1.max_row, len(headers))
    for i, r in enumerate(fd['results'], 1):
        ws1.append([i, r['metric'], r['status'], r['violations'],
                    r['total'], f"{r['percentage']}%", r['threshold']])
        fill = green_fill if r['status'] == "PASS" else red_fill
        for cell in ws1[ws1.max_row]:
            cell.fill = fill; cell.border = thin
    auto_width(ws1)

    # Sheet 2 — Violations
    ws2 = wb.create_sheet("Violations")
    ws2.append(["Metric", "Activity Code", "Activity Name", "Issue"])
    style_header(ws2, 1, 4)
    for r in fd['results']:
        for v in r['details']:
            code = v.get('task_code') or f"{v.get('pred_code','?')}→{v.get('succ_code','?')}"
            name = v.get('task_name') or v.get('pred_name', '')
            ws2.append([r['metric'], code, str(name)[:60], v['issue']])
            for cell in ws2[ws2.max_row]:
                cell.fill = red_fill; cell.border = thin
    auto_width(ws2)

    # Sheet 3 — Engineering
    ws3 = wb.create_sheet("Engineering Checks")
    ws3.append(["Check", "Activity", "Name", "Issue"])
    style_header(ws3, 1, 4)
    for r in fd['eng_results']:
        for v in r['details']:
            code = v.get('task_code') or f"{v.get('pred_code','?')}→{v.get('succ_code','?')}"
            name = v.get('task_name') or v.get('pred_name', '')
            ws3.append([r['check'], code, str(name)[:60], v['issue']])
            for cell in ws3[ws3.max_row]:
                cell.fill = red_fill; cell.border = thin
    auto_width(ws3)

    # Sheet 4 — Activity List
    ws4 = wb.create_sheet("Activity List")
    ws4.append(["Code", "Name", "Type", "Duration (h)", "Float (h)", "Constraint", "Status"])
    style_header(ws4, 1, 7)
    for act in fd['activities'].values():
        ws4.append([act.task_code, act.task_name[:60], act.task_type,
                    act.duration, act.total_float,
                    act.constraint or "None", act.status_code])
    auto_width(ws4)

    output = io.BytesIO(); wb.save(output); output.seek(0)
    return output


# ══════════════════════════════════════════════════════════
# PER-FILE ANALYSIS
# ══════════════════════════════════════════════════════════
file_tab_labels = [f"📁 {fd['filename'][:22]}" for fd in all_file_data]
file_tabs = st.tabs(file_tab_labels)

for file_tab, fd in zip(file_tabs, all_file_data):
    with file_tab:

        results       = fd["results"]
        eng_results   = fd["eng_results"]
        activities    = fd["activities"]
        relationships = fd["relationships"]

        # ── File Header ────────────────────────────────────
        st.markdown(f"""
        <div class="file-header">
            <div style="font-size:28px;">📄</div>
            <div>
                <div class="file-name">{fd['filename']}</div>
                <div class="file-sub">
                    Project: <strong>{fd['proj_name']}</strong> &nbsp;·&nbsp;
                    Data Date: <strong>{fd['data_date'] or 'N/A'}</strong> &nbsp;·&nbsp;
                    Activities: <strong>{len(activities)}</strong> &nbsp;·&nbsp;
                    Relationships: <strong>{len(relationships)}</strong>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # ── Inner Tabs ─────────────────────────────────────
        t1, t2, t3, t4, t5 = st.tabs([
            "🎯 Executive Summary",
            "🟢 Metrics Scorecard",
            "🔍 Violations Explorer",
            "🔧 Engineering",
            "📥 Export"
        ])

        # ══════════════════════════════════════════════════
        # TAB 1 — EXECUTIVE SUMMARY
        # ══════════════════════════════════════════════════
        with t1:
            col_score, col_charts = st.columns([1, 2])

            with col_score:
                st.markdown(f"""
                <div class="score-banner" style="background:{fd['bg_color']};
                     border: 1px solid {fd['score_color']}33;">
                    <div class="score-label" style="color:{fd['score_color']};">
                        OVERALL QUALITY SCORE
                    </div>
                    <div class="score-number" style="color:{fd['score_color']};">
                        {fd['score']}%
                    </div>
                    <div class="score-grade" style="color:{fd['score_color']};">
                        {fd['grade_icon']} {fd['grade']}
                    </div>
                    <hr style="border-color:{fd['score_color']}33; margin:16px 0;">
                    <div style="display:flex; gap:24px; flex-wrap:wrap;">
                        <div>
                            <div style="font-size:28px; font-weight:800;
                                 color:#00b894;">{fd['passed']}</div>
                            <div style="font-size:11px; color:#636e72;
                                 text-transform:uppercase; letter-spacing:1px;">Passing</div>
                        </div>
                        <div>
                            <div style="font-size:28px; font-weight:800;
                                 color:#d63031;">{fd['failed']}</div>
                            <div style="font-size:11px; color:#636e72;
                                 text-transform:uppercase; letter-spacing:1px;">Failing</div>
                        </div>
                        <div>
                            <div style="font-size:28px; font-weight:800;
                                 color:#0d2151;">{len(results)}</div>
                            <div style="font-size:11px; color:#636e72;
                                 text-transform:uppercase; letter-spacing:1px;">Total</div>
                        </div>
                    </div>
                </div>
                """, unsafe_allow_html=True)

            with col_charts:
                c_gauge, c_pie = st.columns(2)

                with c_gauge:
                    gauge = go.Figure(go.Indicator(
                        mode="gauge+number", value=fd['score'],
                        title={"text": "Quality Score",
                               "font": {"color": "#2d3436", "size": 13}},
                        gauge={
                            "axis"   : {"range": [0, 100],
                                        "tickcolor": "#636e72"},
                            "bar"    : {"color": fd['score_color']},
                            "bgcolor": "#f5f7fa",
                            "steps"  : [
                                {"range": [0,  60], "color": "#ffe8e8"},
                                {"range": [60, 80], "color": "#fff8e1"},
                                {"range": [80,100], "color": "#e8f8f0"},
                            ],
                            "threshold": {
                                "line"     : {"color": "#0d2151", "width": 3},
                                "thickness": 0.75, "value": 80
                            }
                        },
                        number={"suffix": "%",
                                "font"  : {"color": fd['score_color'], "size": 36}}
                    ))
                    gauge.update_layout(
                        paper_bgcolor="#ffffff", height=240,
                        margin=dict(t=30, b=10, l=20, r=20)
                    )
                    st.plotly_chart(gauge, use_container_width=True)

                with c_pie:
                    pie = go.Figure(go.Pie(
                        labels=["Passing", "Failing"],
                        values=[fd['passed'], fd['failed']],
                        hole=0.65,
                        marker=dict(
                            colors=["#00b894", "#d63031"],
                            line=dict(color="#ffffff", width=3)
                        ),
                        textinfo="label+value",
                        textfont=dict(size=12, color="#2d3436")
                    ))
                    pie.update_layout(
                        paper_bgcolor="#ffffff", height=240,
                        margin=dict(t=20, b=10, l=10, r=10),
                        showlegend=False,
                        annotations=[dict(
                            text=f"{fd['score']}%",
                            x=0.5, y=0.5, showarrow=False,
                            font=dict(size=26, color=fd['score_color'],
                                      family="Inter")
                        )]
                    )
                    st.plotly_chart(pie, use_container_width=True)

            # Violations bar chart
            st.markdown('<p class="section-title">Violations per Metric</p>',
                        unsafe_allow_html=True)
            bar_names  = [r['metric'].split(' - ')[1] if ' - ' in r['metric']
                          else r['metric'] for r in results]
            bar_values = [r['violations'] for r in results]
            bar_colors = [fd['score_color'] if v == 0 else "#d63031"
                          for v in bar_values]

            fig_bar = go.Figure(go.Bar(
                x=bar_names, y=bar_values,
                marker_color=bar_colors,
                text=bar_values, textposition="outside",
                textfont=dict(size=12, color="#2d3436")
            ))
            fig_bar.update_layout(
                paper_bgcolor="#ffffff", plot_bgcolor="#ffffff",
                height=340, margin=dict(t=20, b=110, l=40, r=20),
                xaxis=dict(tickangle=-35,
                           tickfont=dict(size=10, color="#636e72")),
                yaxis=dict(title="Violations",
                           gridcolor="#f0f3f7",
                           tickfont=dict(color="#636e72")),
                showlegend=False
            )
            st.plotly_chart(fig_bar, use_container_width=True)


        # ══════════════════════════════════════════════════
        # TAB 2 — METRICS SCORECARD (Traffic Light Cards)
        # ══════════════════════════════════════════════════
        with t2:
            st.markdown('<p class="section-title">DCMA 14-Point Metrics</p>',
                        unsafe_allow_html=True)

            # Radar chart
            radar_names  = [r['metric'].split(' - ')[1] if ' - ' in r['metric']
                            else r['metric'] for r in results]
            radar_values = [r['percentage'] for r in results]
            fig_radar = go.Figure(go.Scatterpolar(
                r=radar_values + [radar_values[0]],
                theta=radar_names + [radar_names[0]],
                fill='toself',
                fillcolor="rgba(13,33,81,0.08)",
                line=dict(color="#0d2151", width=2),
            ))
            fig_radar.update_layout(
                paper_bgcolor="#ffffff", height=420,
                margin=dict(t=20, b=20, l=40, r=40),
                polar=dict(
                    bgcolor="#ffffff",
                    radialaxis=dict(visible=True, range=[0, 100],
                                   gridcolor="#dfe6e9",
                                   tickfont=dict(color="#636e72", size=9)),
                    angularaxis=dict(gridcolor="#dfe6e9",
                                    tickfont=dict(color="#2d3436", size=10))
                ),
                showlegend=False
            )
            st.plotly_chart(fig_radar, use_container_width=True)

            st.markdown('<p class="section-title">Metric Cards</p>',
                        unsafe_allow_html=True)

            # Cards in 3 columns
            cols = st.columns(3)
            for i, r in enumerate(results):
                is_pass    = r['status'] == "PASS"
                card_class = "pass" if is_pass else "fail"
                badge      = '<span class="badge-pass">PASS</span>' if is_pass \
                             else '<span class="badge-fail">FAIL</span>'
                stat_color = "#00b894" if is_pass else "#d63031"
                name       = r['metric'].split(' - ')[1] if ' - ' in r['metric'] \
                             else r['metric']

                with cols[i % 3]:
                    st.markdown(f"""
                    <div class="metric-card {card_class}">
                        <div style="display:flex; justify-content:space-between;
                             align-items:flex-start;">
                            <div class="metric-card-title">{name}</div>
                            {badge}
                        </div>
                        <div class="metric-card-stat" style="color:{stat_color};">
                            {r['violations']}
                        </div>
                        <div class="metric-card-sub">
                            violations &nbsp;·&nbsp; {r['percentage']}% of {r['total']}
                        </div>
                        <div style="margin-top:8px; font-size:11px; color:#b2bec3;">
                            Threshold: {r['threshold']}
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

            # Full table below cards
            st.markdown('<p class="section-title">Full Metrics Table</p>',
                        unsafe_allow_html=True)
            df_table = pd.DataFrame([{
                "Metric"     : r['metric'],
                "Status"     : "✅ PASS" if r['status'] == "PASS" else "❌ FAIL",
                "Violations" : r['violations'],
                "Total"      : r['total'],
                "Rate"       : f"{r['percentage']}%",
                "Threshold"  : r['threshold'],
            } for r in results])
            st.dataframe(df_table, use_container_width=True,
                         hide_index=True, height=560)


        # ══════════════════════════════════════════════════
        # TAB 3 — VIOLATIONS EXPLORER
        # ══════════════════════════════════════════════════
        with t3:
            total_v = sum(r['violations'] for r in results)

            col_tv1, col_tv2, col_tv3 = st.columns(3)
            col_tv1.markdown(f"""
            <div class="kpi-box" style="border-top:4px solid #d63031;">
                <div class="kpi-value" style="color:#d63031;">{total_v}</div>
                <div class="kpi-label">Total Violations</div>
            </div>""", unsafe_allow_html=True)
            col_tv2.markdown(f"""
            <div class="kpi-box" style="border-top:4px solid #d63031;">
                <div class="kpi-value" style="color:#d63031;">{fd['failed']}</div>
                <div class="kpi-label">Failed Metrics</div>
            </div>""", unsafe_allow_html=True)
            col_tv3.markdown(f"""
            <div class="kpi-box" style="border-top:4px solid #00b894;">
                <div class="kpi-value" style="color:#00b894;">{fd['passed']}</div>
                <div class="kpi-label">Passed Metrics</div>
            </div>""", unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)

            # Filter
            filter_col, _ = st.columns([2, 3])
            with filter_col:
                filter_opt = st.selectbox(
                    "Filter",
                    ["Show All", "Failed Only", "Passed Only"],
                    key=f"filter_{fd['filename']}"
                )

            # Failed metrics with drill-down
            st.markdown('<p class="section-title">Violation Details</p>',
                        unsafe_allow_html=True)

            shown = 0
            for r in results:
                if filter_opt == "Failed Only"  and r['status'] == "PASS": continue
                if filter_opt == "Passed Only"  and r['status'] == "FAIL": continue

                is_pass = r['status'] == "PASS"
                icon    = "✅" if is_pass else "❌"
                clr     = "#00b894" if is_pass else "#d63031"
                name    = r['metric'].split(' - ')[1] if ' - ' in r['metric'] \
                          else r['metric']

                with st.expander(
                    f"{icon}  {name}  —  "
                    f"{r['violations']} violation(s)  ({r['percentage']}%)",
                    expanded=False
                ):
                    if r['details']:
                        rows = []
                        for v in r['details']:
                            if 'task_code' in v:
                                rows.append({
                                    "Code" : v['task_code'],
                                    "Name" : v.get('task_name', '')[:60],
                                    "Issue": v['issue']
                                })
                            else:
                                rows.append({
                                    "Code" : f"{v.get('pred_code','?')} → {v.get('succ_code','?')}",
                                    "Name" : v.get('pred_name', '')[:50],
                                    "Issue": v['issue']
                                })
                        st.dataframe(pd.DataFrame(rows),
                                     use_container_width=True,
                                     hide_index=True)
                    else:
                        st.success("✅ No violations found for this metric.")
                shown += 1

            if shown == 0:
                st.info("No metrics match the current filter.")


        # ══════════════════════════════════════════════════
        # TAB 4 — ENGINEERING
        # ══════════════════════════════════════════════════
        with t4:
            st.markdown('<p class="section-title">Engineering Validation Checks</p>',
                        unsafe_allow_html=True)

            st.info("""
            Engineering checks validate that activities follow realistic
            EPC construction sequences and energy dependency rules,
            based on industry best practices.
            """)

            eng_pass = sum(1 for r in eng_results if r['status'] == "PASS")
            e1, e2, e3 = st.columns(3)
            e1.markdown(f"""
            <div class="kpi-box" style="border-top:4px solid #0d2151;">
                <div class="kpi-value" style="color:#0d2151;">{len(eng_results)}</div>
                <div class="kpi-label">Total Checks</div>
            </div>""", unsafe_allow_html=True)
            e2.markdown(f"""
            <div class="kpi-box" style="border-top:4px solid #00b894;">
                <div class="kpi-value" style="color:#00b894;">{eng_pass}</div>
                <div class="kpi-label">Passing</div>
            </div>""", unsafe_allow_html=True)
            e3.markdown(f"""
            <div class="kpi-box" style="border-top:4px solid #d63031;">
                <div class="kpi-value" style="color:#d63031;">{len(eng_results)-eng_pass}</div>
                <div class="kpi-label">Failing</div>
            </div>""", unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)

            for r in eng_results:
                is_pass    = r['status'] == "PASS"
                card_class = "pass" if is_pass else "fail"
                badge      = '<span class="badge-pass">PASS</span>' if is_pass \
                             else '<span class="badge-fail">FAIL</span>'
                clr        = "#00b894" if is_pass else "#d63031"

                st.markdown(f"""
                <div class="metric-card {card_class}">
                    <div style="display:flex; justify-content:space-between;">
                        <div class="metric-card-title">{r['check']}</div>
                        {badge}
                    </div>
                    <div class="metric-card-stat" style="color:{clr};">{r['violations']}</div>
                    <div class="metric-card-sub">violations out of {r['total']} checked</div>
                </div>
                """, unsafe_allow_html=True)

                if r['details']:
                    rows = []
                    for v in r['details']:
                        code = v.get('task_code') or \
                               f"{v.get('pred_code','?')}→{v.get('succ_code','?')}"
                        name = v.get('task_name') or v.get('pred_name', '')
                        rows.append({"Activity": code,
                                     "Name"    : str(name)[:55],
                                     "Issue"   : v['issue']})
                    st.dataframe(pd.DataFrame(rows),
                                 use_container_width=True, hide_index=True)
                st.markdown("<br>", unsafe_allow_html=True)

            with st.expander("📖 How Engineering Checks Work"):
                st.markdown("""
                ### 🔩 Mechanical Sequence Check
                Activities must follow: `FOUNDATION → INSTALLATION → TESTING → COMMISSIONING → HANDOVER`
                Any relationship that reverses this order is flagged.

                ### ⚡ Energy Dependency Check
                Energy-dependent activities (testing, commissioning, instruments)
                must have at least one energy source (electrical, power, substation)
                as a direct predecessor.
                """)


        # ══════════════════════════════════════════════════
        # TAB 5 — EXPORT
        # ══════════════════════════════════════════════════
        with t5:
            st.markdown('<p class="section-title">Export Quality Report</p>',
                        unsafe_allow_html=True)

            clean_name = fd['filename'].replace('.xer','').replace('.XER','')

            col_dl1, col_dl2 = st.columns(2)
            with col_dl1:
                st.markdown("""
                <div class="kpi-box" style="text-align:left; margin-bottom:16px;">
                    <div style="font-size:28px; margin-bottom:8px;">📊</div>
                    <div style="font-weight:700; font-size:15px; color:#0d2151;">
                        Excel Report
                    </div>
                    <div style="font-size:12px; color:#636e72; margin-top:4px;">
                        4 sheets: Executive Summary, Violations,
                        Engineering Checks, Full Activity List.
                        Colour-coded cells. Print-ready.
                    </div>
                </div>
                """, unsafe_allow_html=True)
                st.download_button(
                    label=f"⬇️ Download Excel Report",
                    data=generate_excel(fd),
                    file_name=f"JESA_ScheduleQuality_{clean_name}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

            with col_dl2:
                st.markdown("""
                <div class="kpi-box" style="text-align:left; margin-bottom:16px;">
                    <div style="font-size:28px; margin-bottom:8px;">📄</div>
                    <div style="font-weight:700; font-size:15px; color:#0d2151;">
                        CSV Summary
                    </div>
                    <div style="font-size:12px; color:#636e72; margin-top:4px;">
                        Lightweight flat-file summary of all
                        metric results. Compatible with Excel,
                        Power BI, and any reporting tool.
                    </div>
                </div>
                """, unsafe_allow_html=True)
                csv_rows = [{
                    "Metric"    : r['metric'],
                    "Status"    : r['status'],
                    "Violations": r['violations'],
                    "Total"     : r['total'],
                    "Rate %"    : r['percentage'],
                    "Threshold" : r['threshold']
                } for r in results]
                st.download_button(
                    label=f"⬇️ Download CSV Summary",
                    data=pd.DataFrame(csv_rows).to_csv(index=False),
                    file_name=f"JESA_ScheduleQuality_{clean_name}.csv",
                    mime="text/csv",
                    use_container_width=True
                )

            st.markdown('<p class="section-title">Report Preview</p>',
                        unsafe_allow_html=True)
            st.dataframe(pd.DataFrame([{
                "Metric"     : r['metric'],
                "Status"     : "✅ PASS" if r['status'] == "PASS" else "❌ FAIL",
                "Violations" : r['violations'],
                "Rate"       : f"{r['percentage']}%",
                "Threshold"  : r['threshold'],
            } for r in results]), use_container_width=True, hide_index=True)
